import pandas as pd 
from flask import request, jsonify
from flask_jwt_extended import jwt_required
from datetime import datetime
from openpyxl import load_workbook
from app.utils import Error, Success
import calendar
from app.models import Attendance , Deducation , PayrollDeducation , Payroll
from app.extencions import db
import uuid
from sqlalchemy import and_, or_
from datetime import datetime
# def calculate_work_hours_and_overtime(time_in, time_out):
#     if time_in and time_out:
#         # Chuyển đổi time_in và time_out thành datetime
#         today = datetime.today()
#         time_in_datetime = datetime.combine(today, time_in)
#         time_out_datetime = datetime.combine(today, time_out)

#         # Tính toán thời gian làm việc
#         work_duration = time_out_datetime - time_in_datetime
#         work_hours = work_duration.total_seconds() / 3600  # Chuyển đổi giây thành giờ

#         # Tính overtime
#         total_overtime = round(work_hours / 8, 2) - 1 if work_hours > 8 else None
#         return round(work_hours / 8, 2), total_overtime
#     return None, None
def calculate_work_hours_and_overtime(time_in, time_out):
    from datetime import datetime, time as dt_time

    def to_time_str(t):
        if isinstance(t, dt_time):
            return t.strftime("%H:%M:%S")
        return t

    if time_in and time_out:
        t_in_str = to_time_str(time_in)
        t_out_str = to_time_str(time_out)
        t_in = datetime.strptime(t_in_str, "%H:%M:%S")
        t_out = datetime.strptime(t_out_str, "%H:%M:%S")
        work_hours = (t_out - t_in).total_seconds() / 3600
        overtime_hours = work_hours - 8 if work_hours > 8 else 0
        return round(work_hours, 2), round(overtime_hours, 2) if overtime_hours >= 1 else None
    return None, None
class AttendanceService:   
    @staticmethod 
    def  upload_attendance_file(file):
        try:   
            # Đọc file Excel mà không ép kiểu dữ liệu
            df = pd.read_excel(file, engine='openpyxl',header=1, dtype=str)
            df.columns = [str(col).strip().replace(' ', '_') for col in df.columns]
            
            # Thay thế NaN bằng chuỗi rỗng
            df = df.fillna("")
            # Chuyển DataFrame thành danh sách các dictionary
            wb = load_workbook(file)
            sheet = wb.active
            
            month_year_cell = sheet.cell(row=1, column=2).value  # Giá trị ô B1
            if not month_year_cell:
                return Error(
                    message="Không tìm thấy thông tin tháng/năm trong ô B1.",
                    status=400
                ).to_json(), 400
            
            month = month_year_cell.month if isinstance(month_year_cell, datetime) else None
            year = month_year_cell.year if isinstance(month_year_cell, datetime) else None
    
            _, number_of_days = calendar.monthrange(year, month)
            valid_days = [str(day) for day in range(1, number_of_days + 1)]
            # Tạo dictionary để lưu comment
            comments = {}

            for row in sheet.iter_rows():
                for cell in row:
                    if cell.comment:  # Nếu ô có comment
                        comments[(cell.row, cell.column)] = cell.comment.text

            # Thêm comment vào DataFrame
            for (row, col), comment in comments.items():
                try:
                    # Điều chỉnh chỉ số hàng để khớp với DataFrame
                    df_row_index = row - 3  # Điều chỉnh theo cấu trúc Excel của bạn
                    
                    # Debug: In thông tin để kiểm tra
                    print(f"Excel Row: {row}, Excel Column: {col}, DataFrame Index: {df_row_index}")
                    print(f"Comment: {comment}")
                    
                    if df_row_index >= 0 and df_row_index < len(df):  # Kiểm tra chỉ số hàng hợp lệ
                        if col - 1 < len(df.columns):
                            col_name = df.columns[col - 1]  # Lấy tên cột từ DataFrame
                            if col_name in df.columns:  # Kiểm tra xem cột có tồn tại trong DataFrame
                                print(f"Adding comment to df.at[{df_row_index}, '{col_name}_comment']")
                                df.at[df_row_index, f"{col_name}_comment"] = comment  # Thêm comment vào DataFrame
                            else:
                                print(f"Column name '{col_name}' not found in DataFrame columns")
                        else:
                            print(f"Column index {col-1} out of range for DataFrame columns")
                    else:
                        print(f"Invalid row index: {df_row_index} for Excel row: {row}")
                except Exception as e:
                    print(f"Error processing comment at row {row}, column {col}: {str(e)}")

          
            attendance_records = []

            for _, row in df.iterrows():
                for day in valid_days:  # Chỉ xử lý các cột ngày hợp lệ
                    if day in row:
                        # Lấy giá trị comment và đảm bảo nó là chuỗi
                        comment = row.get(f"{day}_comment", "")
                        if not comment:  # Nếu không có comment
                          # Debug: In thông báo nếu không có comment
                            time_in = None
                            time_out = None
                        else:
                            if not isinstance(comment, str):  # Kiểm tra nếu không phải chuỗi
                                comment = str(comment) if comment else ""

                            # Tách time_in và time_out từ comment
                            time_in = None
                            time_out = None
                            if "\n" in comment:
                                parts = comment.split("\n")
                                # Debug: In các phần sau khi tách
                                if len(parts) >= 3:  # Kiểm tra nếu có đủ phần tử
                                    time_in = parts[1].strip()  # Lấy dòng thứ 2 (giờ vào)
                                    time_out = parts[2].strip()  # Lấy dòng thứ 3 (giờ ra)
                                elif len(parts) == 2:  # Nếu chỉ có ngày và giờ vào
                                    time_in = parts[1].strip()
                                    time_out = None
                                else:  # Nếu chỉ có ngày
                                    time_in = None
                                    time_out = None
                            else:
                              # Debug: In thông báo nếu không có '\n'
                                time_in = None
                                time_out = None

                        # Tạo bản ghi
                        if time_in and time_out:
                            
                            record = { 
                                "attendance_date": f"{year}-{month:02d}-{int(day):02d}",  # Tạo ngày từ cột
                                "time_in": time_in,
                                "time_out": time_out,
                                "total_overtime": round(round((datetime.strptime(time_out,"%H:%M:%S") - datetime.strptime(time_in,"%H:%M:%S")).total_seconds()/3600) / 8 , 2) - 1 if time_in and time_out and round(round((datetime.strptime(time_out,"%H:%M:%S") - datetime.strptime(time_in,"%H:%M:%S")).total_seconds()/3600) / 8 , 2) > 1 else None,  # Giả sử không có làm thêm giờ
                                "note": comment,
                                "employee_id": row.get("Mã_nhân_viên"),
                                "attendance_status_id": 1 if time_in and time_out and round(round((datetime.strptime(time_out,"%H:%M:%S") - datetime.strptime(time_in,"%H:%M:%S")).total_seconds()/3600) / 8 , 2) < 1 else None,  # Giả sử trạng thái là 1
                            }
                        attendance_records.append(record)
            # Chuyển DataFrame thành danh sách các dictionary
            data = df.to_dict(orient='records')
            employee_ids = set(record["employee_id"] for record in attendance_records)

            # Xóa toàn bộ attendance của nhân viên này trong tháng/năm đang upload
            for emp_id in employee_ids:
                db.session.query(Attendance).filter(
                    Attendance.employee_id == emp_id,
                    db.extract('month', Attendance.attendance_date) == month,
                    db.extract('year', Attendance.attendance_date) == year
                ).delete()
            db.session.commit()
            # Lưu vào cơ sở dữ liệu
            unique_keys = set()
            filtered_records = []
            for record in attendance_records:
                key = (record["employee_id"], record["attendance_date"])
                if key not in unique_keys:
                    filtered_records.append(record)
                    unique_keys.add(key)
            attendance_records = filtered_records
            employee_ids = set(record["employee_id"] for record in attendance_records)

            # Tạo danh sách tất cả các ngày làm việc trong tháng (thứ 2-6)
            all_dates = []
            for day in range(1, number_of_days + 1):
                d = datetime(year, month, day)
                if d.weekday() < 5:  # 0-4 là thứ 2-6
                    all_dates.append(d.strftime("%Y-%m-%d"))

            for emp_id in employee_ids:
                # Lấy các ngày đã có chấm công của nhân viên này
                emp_dates = set(
                    record["attendance_date"] for record in attendance_records if record["employee_id"] == emp_id
                )
                # Thêm bản ghi nghỉ không phép cho ngày chưa có chấm công
                for d in all_dates:
                    if d not in emp_dates:
                        attendance_records.append({
                            "attendance_date": d,
                            "time_in": None,
                            "time_out": None,
                            "note": "Nghỉ không phép",
                            "employee_id": emp_id,
                            "attendance_status_id": 3,  # hoặc 1 nếu bạn muốn đánh dấu là nghỉ không phép
                        })

            for record in attendance_records:
                attendance = Attendance(
                    id=str(uuid.uuid4()), 
                    attendance_date=record["attendance_date"],
                    time_in=record["time_in"],
                    time_out=record["time_out"],
                    note=record["note"],
                    employee_id=record["employee_id"],
                    attendance_status_id=record['attendance_status_id'],  # Giả sử trạng thái là 1
                )
                db.session.add(attendance)
                db.session.commit()
                work_hours, _ = calculate_work_hours_and_overtime(attendance.time_in, attendance.time_out)
                if work_hours is not None and work_hours < 1:
                    # Tìm hoặc tạo loại khấu trừ
                    deducation = Deducation.query.filter_by(type_deducation="Kỷ luật", name_deducation="Khấu trừ đi muộn").first()
                    if not deducation:
                        deducation = Deducation(
                            id=str(uuid.uuid4()),
                            type_deducation="Kỷ luật",
                            name_deducation="Khấu trừ đi muộn",
                            money=100000,  # Số tiền khấu trừ, điều chỉnh theo quy định
                            date_deducation=datetime.now().date()
                        )
                        db.session.add(deducation)
                        db.session.commit()
                    # Tìm payroll của nhân viên trong tháng đó
                    month = attendance.attendance_date.month  # Kết quả: 4 (kiểu int)
                    year = attendance.attendance_date.year 
                    payroll = Payroll.query.filter(
                        Payroll.employee_id == attendance.employee_id,
                        db.extract('month', Payroll.month) == month,
                        Payroll.year == year
                    ).first()
                    if payroll:
    # Kiểm tra đã có phiếu phạt cho ngày này chưa
                        existing = PayrollDeducation.query.filter_by(
                            payroll_id=payroll.id,
                            deducation_id=deducation.id,
                            attendance_date=attendance.attendance_date
                        ).first()
                        if not existing:
                            payroll_deducation = PayrollDeducation(
                                id=str(uuid.uuid4()),
                                payroll_id=payroll.id,
                                deducation_id=deducation.id,
                                attendance_date=attendance.attendance_date
                            )
                            db.session.add(payroll_deducation)
                            db.session.commit()
            return Success(
                message="Upload file thành công",
                status=200,
                payload={},  # Trả về dữ liệu gốc
            ).to_json(), 200
        except Exception as e:
            print(e)
            return Error(
                message=str(e),
                status=400
            ).to_json(), 400
    

    @staticmethod
    def get_attendance_by_employee_id(employee_id):
        try:
            # Lấy tất cả bản ghi Attendance cho employee_id
            attendance_records = Attendance.query.filter_by(employee_id=employee_id).all()
            
            # Chuyển đổi các bản ghi thành danh sách dictionary
            attendance_list = [record.to_dict() for record in attendance_records]
            
            return Success(
                message="Lấy danh sách chấm công thành công",
                status=200,
                payload=attendance_list,
            ).to_json(), 200
        except Exception as e:
            return Error(
                message=str(e),
                status=400
            ).to_json(), 400
        
    @staticmethod
    def get_attendance_by_month_year(month, year):
        try:
            # Lấy tất cả bản ghi Attendance cho tháng và năm
            attendance_records = Attendance.query.filter(
                db.extract('month', Attendance.attendance_date) == month,
                db.extract('year', Attendance.attendance_date) == year
            ).all()
            
            # Chuyển đổi các bản ghi thành danh sách dictionary
            attendance_list = [record.to_dict() for record in attendance_records]
            
            return Success(
                message="Lấy danh sách chấm công thành công",
                status=200,
                payload=attendance_list,
            ).to_json(), 200
        except Exception as e:
            return Error(
                message=str(e),
                status=400
            ).to_json(), 400
        
    @staticmethod
    def get_attendance_by_employee_id_and_month_year(id, year, month):
        try:
            # Lấy tất cả bản ghi Attendance cho employee_id, tháng và năm
            attendance_records = Attendance.query.filter(
                and_(
                    Attendance.employee_id == id,
                    db.extract('month', Attendance.attendance_date) == month,
                    db.extract('year', Attendance.attendance_date) == year
                )
            ).all()
            
            # Chuyển đổi các bản ghi thành danh sách dictionary
            # attendance_list = [record.to_dict() for record in attendance_records]
            
        # Chuyển đổi các bản ghi thành danh sách dictionary
            result_data = []
            for record in attendance_records:
                work_hours, total_overtime = calculate_work_hours_and_overtime(
                        record.time_in.strftime('%H:%M:%S') if record.time_in else None,
                        record.time_out.strftime('%H:%M:%S') if record.time_out else None
                    )
                record_dict = {
                        "id": record.id,
                        "attendance_date": record.attendance_date.strftime('%Y-%m-%d'),
                        "time_in": record.time_in.strftime('%H:%M:%S') if record.time_in else None,
                        "time_out": record.time_out.strftime('%H:%M:%S') if record.time_out else None,
                        "total_overtime": total_overtime,
                        "note": record.note,
                        "employee_id": record.employee_id,
                        "attendance_status_id": record.attendance_status_id,
                        "workHours": work_hours
                    }
                result_data.append(record_dict)
            return Success(
                message="Lấy danh sách chấm công thành công",
                status=200,
                payload=result_data,
            ).to_json(), 200
        except Exception as e:
            return Error(
                message=str(e),
                status=400
            ).to_json(), 400

    @staticmethod
    def get_attendance_by_employee_id_and_date(employee_id, date):
        try:
            # Lấy tất cả bản ghi Attendance cho employee_id và ngày
            attendance_records = Attendance.query.filter(
                Attendance.employee_id == employee_id,
                db.func.date(Attendance.attendance_date) == date
            ).all()
            
            # Chuyển đổi các bản ghi thành danh sách dictionary
            attendance_list = [record.to_dict() for record in attendance_records]
            
            return Success(
                message="Lấy danh sách chấm công thành công",
                status=200,
                payload=attendance_list,
            ).to_json(), 200
        except Exception as e:
            return Error(
                message=str(e),
                status=400
            ).to_json(), 400
    
    @staticmethod
    def add_attendance(attendance_data):
        try:
            # Tạo bản ghi Attendance mới
            attendance = Attendance(
                id=str(uuid.uuid4()),  # Tạo ID ngẫu nhiên
                attendance_date=attendance_data["attendance_date"],
                time_in=attendance_data["time_in"],
                time_out=attendance_data["time_out"],
                note=attendance_data["note"] if attendance_data["note"] else None,
                total_overtime=attendance_data["total_overtime"] if attendance_data["total_overtime"] else None,
                employee_id=attendance_data["employee_id"],
                attendance_status_id=attendance_data['attendance_status_id'] if attendance_data['attendance_status_id'] else None,  # Giả sử trạng thái là 1
            )
            
            db.session.add(attendance)
            db.session.commit()
            
            return Success(
                message="Thêm chấm công thành công",
                status=200,
                # payload={attendance.to_dict()}
            ).to_json(), 200
        except Exception as e:
            print(e)
            return Error(
                message=str(e),
                status=400
            ).to_json(), 400
        
    @staticmethod 
    def delete_attendance(attendance_id):
        try:
            # Xóa bản ghi Attendance theo ID
            attendance = Attendance.query.get(attendance_id)
            if not attendance:
                return Error(
                    message="Không tìm thấy bản ghi chấm công",
                    status=404
                ).to_json(), 404
            
            db.session.delete(attendance)
            db.session.commit()
            
            return Success(
                message="Xóa chấm công thành công",
                status=200,
                payload={}
            ).to_json(), 200
        except Exception as e:
            print(e)
            return Error(
                message=str(e),
                status=400
            ).to_json(), 400
    
    @staticmethod
    def update_attendance( attendance_data):
        try:
            # Cập nhật bản ghi Attendance theo ID
            attendance = Attendance.query.get(attendance_data['id'])
            if not attendance:
                return Error(
                    message="Không tìm thấy bản ghi chấm công",
                    status=404
                ).to_json(), 404
            
            # Cập nhật các trường cần thiết
            attendance.attendance_date = attendance_data["attendance_date"]
            attendance.time_in = attendance_data["time_in"]
            attendance.time_out = attendance_data["time_out"]
            attendance.note = attendance_data["note"] if attendance_data["note"] else None
            attendance.total_overtime = attendance_data["total_overtime"] if attendance_data["total_overtime"] else None
            attendance.employee_id = attendance_data["employee_id"]
            attendance.attendance_status_id = attendance_data['attendance_status_id'] if attendance_data['attendance_status_id'] else None  # Giả sử trạng thái là 1
            
            db.session.commit()
            
            return Success(
                message="Cập nhật chấm công thành công",
                status=200,
                payload={}
            ).to_json(), 200
        except Exception as e:
            print(e)
            return Error(
                message=str(e),
                status=400
            ).to_json(), 400